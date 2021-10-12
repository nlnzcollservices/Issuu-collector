import os
import re
import time
import sys
import wget
import requests
import urllib
import shutil
import dateparser
import pyautogui
import keyboard
from bs4 import BeautifulSoup as bs
from datetime import datetime as dt
from time import mktime, sleep
from selenium import webdriver
from sys import platform
from issuu_dict import issuu_dict
from my_settings import sip_folder, temp_folder,report_folder, template_folder,logging, rosetta_folder
sys.path.insert(0,r'Y:\ndha\pre-deposit_prod\LD_working\podcasts\scripts')
from alma_tools import AlmaTools
# sys.path.insert(0,r'Y:\ndha\pre-deposit_prod\LD_working\waterford\scripts')
# import last_representation_getter
import send_email
from openpyxl import Workbook, load_workbook

sprdsht_path = r"Y:\ndha\pre-deposit_prod\LD_working\issuu_main\assets\metadata\Copy of Issuu_titles_sent_Svetlana_Oct_2021.xlsx"
#['_Workbook__write_only', '__class__', '__contains__', '__delattr__', '__delitem__', '__dict__', '__dir__', '__doc__', '__eq__', '__format__', '__ge__', '__getattribute__', '__getitem__', '__gt__', '__hash__', '__init__', '__init_subclass__', '__iter__', '__le__', '__lt__', '__module__', '__ne__', '__new__', '__reduce__', '__reduce_ex__', '__repr__', '__setattr__', '__sizeof__', '__str__', '__subclasshook__', '__weakref__', '_active_sheet_index', '_add_sheet', '_alignments', '_borders', '_cell_styles', '_colors', '_data_only', '_date_formats', '_differential_styles', '_duplicate_name', '_epoch', '_external_links', '_fills', '_fonts', '_named_styles', '_number_formats', '_pivots', '_protections', '_read_only', '_setup_styles', '_sheets', '_table_styles', '_timedelta_formats', 'active', 'add_named_range', 'add_named_style', 'calculation', 'chartsheets', 'close', 'code_name', 'copy_worksheet', 'create_chartsheet', 'create_named_range', 'create_sheet', 'data_only', 'defined_names', 'encoding', 'epoch', 'excel_base_date', 'get_index', 'get_named_range', 'get_named_ranges', 'get_sheet_by_name', 'get_sheet_names', 'index', 'is_template', 'iso_dates', 'loaded_theme', 'mime_type', 'move_sheet', 'named_styles', 'path', 'properties', 'read_only', 'rels', 'remove', 'remove_named_range', 'remove_sheet', 'save', 'security', 'shared_strings', 'sheetnames', 'style_names', 'template', 'vba_archive', 'views', 'worksheets', 'write_only']
#['__add__', '__class__', '__class_getitem__', '__contains__', '__delattr__', '__dir__', '__doc__', '__eq__', '__format__', '__ge__', '__getattribute__', '__getitem__', '__getnewargs__', '__gt__', '__hash__', '__init__', '__init_subclass__', '__iter__', '__le__', '__len__', '__lt__', '__mul__', '__ne__', '__new__', '__reduce__', '__reduce_ex__', '__repr__', '__rmul__', '__setattr__', '__sizeof__', '__str__', '__subclasshook__', 'count', 'index']
#print(issuu_dict["Geraldine news"].keys())
{'publisher':"", 'web_title':"", 'mms_id':"", 'holding_id':"", 'po_line':"", 'pattern':"", 'days':"", 'url':"", 'username':""}

def routine():
	my_alma = AlmaTools("prod")
	wb = load_workbook(sprdsht_path)
	ws = wb.get_sheet_by_name("sent_Svetlana_Oct_2021")
	print(dir(ws))
	for i in range(ws.max_row):
		if i>1:
			print("__________________________________")
			# for col in ["A","B","C","D","E","F","G","H","I","J","K","L"]:
			# 	print(ws["{}{}".format(col,i+1)].value)
			#issuu_dict["D{}".format(i+1)] ={'publisher':"", 'web_title':"", 'mms_id':"", 'holding_id':"", 'po_line':"", 'pattern':"", 'days':"", 'url':"", 'username':""}
			mms = ws["C{}".format(i+1)].value
			if mms:
				my_alma.get_holdings(mms)
				#print(my_alma.xml_response_data)
				holding_id = re.findall(r'<holding_id>(.*?)</holding_id>',my_alma.xml_response_data)[0]
				#print(holding_id)
				publisher = ws["G{}".format(i+1)].value
				url = ws["B{}".format(i+1)].value
				web_title_r= requests.get(url)
				soup = bs(web_title_r.text,"lxml")
				web_title = soup.find_all("h1")[0].text
				username = url.split("/")[-1]
				po_line = ws["J{}".format(i+1)].value
				issuu_dict[ws["D{}".format(i+1)].value] ={'publisher':publisher, 'web_title':web_title, 'mms_id':mms, 'holding_id':holding_id, 'po_line':po_line, 'pattern':"", 'days':"", 'url':url, 'username':username}
	print(issuu_dict)


if __name__ == "__main__":
	routine()
