import os
import re
import time
import sys
import requests
import urllib
import shutil
import dateparser
import gspread
from bs4 import BeautifulSoup as bs
from datetime import datetime as dt
from time import mktime, sleep
from sys import platform
from issuu_dict import issuu_dict
from my_settings import sip_folder, temp_folder,report_folder, template_folder,logging, rosetta_folder
sys.path.insert(0,r'Y:\ndha\pre-deposit_prod\LD_working\podcasts\scripts')
from alma_tools import AlmaTools
from oauth2client import file
from openpyxl import Workbook, load_workbook

sprdsht_path = r"Y:\ndha\pre-deposit_prod\LD_working\issuu_main\assets\metadata\Adding and checking Issuu titles list.xlsx"
google_cred_path = r"H:\secrets_and_credentials\client_secrets.json"
google_sprsh_key = r"1TqO0IEdNiyJRPk-IzwTzqx3F52mqgyp_YkFpc2UkbfA"
store = file.Storage(google_cred_path)
creds = store.get()
if creds.access_token_expired:
	creds.refresh(httplib2.Http())

c = gspread.authorize(creds)
gs = c.open_by_key(google_sprsh_key)
#change if the name or id of the worksheet is different
ws = gs.get_worksheet(0)

issuu_dict= {}
my_alma = AlmaTools("prod")

def google_spreadsheet_routine():
	#print(ws.row_count)
	for i in range( ws.row_count):
		row_number = i
		if i in list(range(166,172)):
			#print(ws.row_values(row_number))
			print("__________________________________")
			try:
				print(ws.acell("B{}".format(row_number)).value)
				mms = ws.acell("D{}".format(row_number)).value.rstrip(" ")
				print(mms)

				my_alma.get_holdings(mms)
				#print(my_alma.xml_response_data)
				holding_id = re.findall(r'<holding_id>(.*?)</holding_id>',my_alma.xml_response_data)[0]
				#print(holding_id)
				publisher = ws.acell("C{}".format(row_number)).value.rstrip(" ").rstrip(".")
				try:
					url = ws.acell("A{}".format(row_number)).value.rstrip(" ")
					print(url)
					web_title_r= requests.get(url, verify = False)

					soup = bs(web_title_r.text,"lxml")
					print("here")
					web_title = soup.find_all("h1")[0].text
					print("here2")
					username = url.split("/")[-1].rstrip(" ")
				except:
					url = None
					web_title = None
					username = None
				po_line = ws.acell("E{}".format(row_number)).value.rstrip(" ")
				try:
					my_alma.get_po_line(po_line)
					days = re.findall(r'<subscription_interval>(.*?)</subscription_interval',my_alma.xml_response_data )[0]
				except Exception as e:
					print(str(e))
					days = None
				access = str(ws.acell("F{}".format(row_number)).value).rstrip(" ")
				
				issuu_dict[ws.acell("B{}".format(row_number)).value] ={'publisher':publisher, 'web_title':web_title, 'mms_id':mms, 'holding_id':holding_id, 'po_line':po_line, 'pattern':"", 'days':days, 'url':url, 'username':username, 'access':access}
				for el in issuu_dict:
					print(f'"{el}":',issuu_dict[el],",")
			
			except Exception as e:
				print(str(e))
	print(issuu_dict)



def routine():
	

	wb = load_workbook(sprdsht_path)
	ws = wb.get_sheet_by_name("adding_checking_titles_list")#("sent_Svetlana_Oct_2021")#
	print(dir(ws))

	for i in range(ws.max_row):
		row_number = i+1
		if row_number in range(158,163):
			po_line = None
			days= None
			print("__________________________________")
			# for col in ["A","B","C","D","E","F","G","H","I","J","K","L"]:
			# 	print(ws["{}{}".format(col,row_number)].value)
			#issuu_dict["D{}".format(row_number)] ={'publisher':"", 'web_title':"", 'mms_id':"", 'holding_id':"", 'po_line':"", 'pattern':"", 'days':"", 'url':"", 'username':""}
			try:
				print(ws["B{}".format(row_number)].value)
				mms = ws["D{}".format(row_number)].value.rstrip(" ")
				print(mms)

				my_alma.get_holdings(mms)
				#print(my_alma.xml_response_data)
				holding_id = re.findall(r'<holding_id>(.*?)</holding_id>',my_alma.xml_response_data)[0]
				#print(holding_id)
				publisher = ws["C{}".format(row_number)].value.rstrip(" ").rstrip(".")
				try:
					url = ws["A{}".format(row_number)].value.rstrip(" ")
					print(url)
					web_title_r= requests.get(url)
					soup = bs(web_title_r.text,"lxml")
					print("here")
					web_title = soup.find_all("h1")[0].text
					print("here2")
					username = url.split("/")[-1].rstrip(" ")
				except:
					url = None
					web_title = None
					username = None
				po_line = ws["E{}".format(row_number)].value.rstrip(" ")
				try:
					my_alma.get_po_line(po_line)
					days = re.findall(r'<subscription_interval>(.*?)</subscription_interval',my_alma.xml_response_data )[0]
				except Exception as e:
					print(str(e))
					days = None
				access = str(ws["F{}".format(row_number)].value).rstrip(" ")
				
				issuu_dict[ws["B{}".format(row_number)].value] ={'publisher':publisher, 'web_title':web_title, 'mms_id':mms, 'holding_id':holding_id, 'po_line':po_line, 'pattern':"", 'days':days, 'url':url, 'username':username, 'access':access}
				for el in issuu_dict:
					print(f'"{el}":',issuu_dict[el],",")
			
			except Exception as e:
				print(str(e))
	print(issuu_dict)


if __name__ == "__main__":
	#routine()
	google_spreadsheet_routine()
